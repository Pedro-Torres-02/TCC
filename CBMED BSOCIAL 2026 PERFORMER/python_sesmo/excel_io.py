"""Leitura e escrita das planilhas Excel (openpyxl)."""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook

import config


@dataclass
class LinhaFila:
    """Uma linha completa da planilha de Fila (Plan1)."""
    excel_row: int       # número da linha no Excel (1-based; dados começam em 2)
    colaborador: str     # col A (índice 0)
    cpf: str             # col B (índice 1)
    cnpj: str            # col C (índice 2)
    data_inconsistencia: str  # col D (índice 3)
    inconsistencia: str  # col E (índice 4)
    status: str          # col F (índice 5)
    mensagem: str        # col G (índice 6)
    matricula: str       # col H (índice 7)
    situacao: str        # col I (índice 8)


def ler_fila(caminho: str, sheet: str = config.SHEET_FILA) -> list[LinhaFila]:
    """Lê a planilha de fila completa (pula o cabeçalho na linha 1)."""
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb[sheet]
    linhas: list[LinhaFila] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def col(idx): return _txt(row[idx]) if len(row) > idx else ""
        if not col(0) and not col(1):
            continue  # linha vazia
        linhas.append(LinhaFila(
            excel_row=i,
            colaborador=col(0),
            cpf=col(1),
            cnpj=col(2),
            data_inconsistencia=col(3),
            inconsistencia=col(4),
            status=col(5),
            mensagem=col(6),
            matricula=col(7),
            situacao=col(8),
        ))
    wb.close()
    return linhas


def escrever_importados_sesmo(caminho: str, resultados: list[tuple[str, str]],
                              sheet: str = config.SHEET_IMPORTADOS_SESMO) -> None:
    """Grava todos os resultados extraídos do Sesmo na Planilha Importados Sesmo.

    resultados: lista de (nome_colaborador, aviso) da tabela do Sesmo.
    Col A = Nome do Colaborador, Col B = Aviso, a partir da linha 2.
    """
    if not resultados:
        return
    wb = load_workbook(caminho)
    ws = wb[sheet]
    for i, (nome, aviso) in enumerate(resultados, start=2):
        ws[f"A{i}"] = nome
        ws[f"B{i}"] = aviso
    wb.save(caminho)
    wb.close()


def escrever_revisao(caminho: str, marcacoes: list[tuple[int, str]],
                     sheet: str = config.SHEET_FILA) -> None:
    """Escreve F='Revisão' e G=<aviso> nas linhas com erro.

    marcacoes: lista de (excel_row, aviso_sesmo) — somente erros, já filtrado pelo cruzar().
    """
    if not marcacoes:
        return
    wb = load_workbook(caminho)
    ws = wb[sheet]
    for excel_row, aviso in marcacoes:
        ws[f"F{excel_row}"] = "Revisão"
        ws[f"G{excel_row}"] = aviso
    wb.save(caminho)
    wb.close()


def _txt(v) -> str:
    return "" if v is None else str(v).strip()
