"""Envio dos resultados ao Dash.

Espelha o 'Envio Dash.xaml': para cada linha da planilha de Fila monta um JSON
e envia. No UiPath isso era feito chamando uipath_bridge.envio_dash(json).

Dois modos (configuráveis no .env):
  - DASH_BRIDGE_PATH: importa o uipath_bridge.py existente e chama envio_dash(json)
  - DASH_ENDPOINT: faz POST do JSON direto para o endpoint

Estrutura do JSON (idêntica ao UiPath):
  {
    "status": <col5>,
    "mensagem": <col6>,
    "info_referencia": <cpf/col1>,
    "extra_metadata": {
      "fila": <"2220"|"2240">,
      "nome": <col0>, "cpf": <col1>, "cnpj": <col2>,
      "data_inconsistencia": <col3>, "inconsistencia": <col4>,
      "matricula": <col7>, "situacao": <col8>,
      "data_exec": <"dd/MM/yyyy - HH:mm:ss">
    }
  }
"""
from __future__ import annotations

import importlib.util
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import config


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def montar_payloads(caminho_fila: str, fila_nome: str,
                    sheet: str = config.SHEET_FILA) -> list[dict]:
    """Lê a planilha de fila e monta um payload (dict) por linha."""
    data_exec = datetime.now().strftime("%d/%m/%Y - %H:%M:%S")
    wb = load_workbook(caminho_fila, data_only=True, read_only=True)
    ws = wb[sheet]
    payloads: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        c = [_txt(row[i]) if len(row) > i else "" for i in range(9)]
        if not any(c):
            continue
        extra = {
            "fila": fila_nome,
            "nome": c[0], "cpf": c[1], "cnpj": c[2],
            "data_inconsistencia": c[3], "inconsistencia": c[4],
            "matricula": c[7], "situacao": c[8],
            "data_exec": data_exec,
        }
        payloads.append({
            "status": c[5],
            "mensagem": c[6],
            "info_referencia": c[1],
            "extra_metadata": extra,
        })
    wb.close()
    return payloads


def _carregar_bridge(bridge_dir: str):
    """Importa uipath_bridge.py do diretório configurado."""
    bridge_file = Path(bridge_dir) / "uipath_bridge.py"
    if not bridge_file.exists():
        raise FileNotFoundError(f"uipath_bridge.py não encontrado em {bridge_dir}")
    spec = importlib.util.spec_from_file_location("uipath_bridge", bridge_file)
    mod = importlib.util.module_from_spec(spec)
    sys.modules["uipath_bridge"] = mod
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


def enviar(payloads: list[dict]) -> None:
    """Envia cada payload ao dash, conforme o modo configurado."""
    if config.DASH_BRIDGE_PATH:
        bridge = _carregar_bridge(config.DASH_BRIDGE_PATH)
        for p in payloads:
            bridge.envio_dash(json.dumps(p, ensure_ascii=False))
    elif config.DASH_ENDPOINT:
        import requests
        for p in payloads:
            resp = requests.post(config.DASH_ENDPOINT, json=p, timeout=30)
            resp.raise_for_status()
    else:
        raise RuntimeError(
            "Nenhum destino de dash configurado (DASH_BRIDGE_PATH ou DASH_ENDPOINT no .env)"
        )


def enviar_fila(caminho_fila: str, fila_nome: str) -> int:
    """Monta e envia os payloads de uma fila. Retorna a quantidade enviada."""
    payloads = montar_payloads(caminho_fila, fila_nome)
    enviar(payloads)
    return len(payloads)
